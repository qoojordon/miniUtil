import sys
from typing import Tuple
print("start procedure...")
import openpyxl

RESULT_SHEET_NAMES=("Result_Author_Keyword", "Result_Keyword_Plus")
SUPPORTED_FIELDS=('Article Title', 'Author Keywords', 'Keywords Plus')
FEQ_FIELDS      =                 ('Author Keywords', 'Keywords Plus')
g_field_2_idx = {}
g_f_2_kw_map_feq = {}
g_warnings = []

def error_exit(msg: str):
    print(f"[ERROR] {msg}")
    sys.exit(1)

def pwarn(msg: str):
    suppres_upper_bound = 10
    if len(g_warnings) < suppres_upper_bound:
        print(f"[WARN] {msg}")
    g_warnings.append(msg)
    if len(g_warnings) == suppres_upper_bound:
        print(f"[WARN] hit over {suppres_upper_bound} warnings, suppress later warnings")


def verify_head_row_and_build_global_var(row : Tuple):
    for idx, cell in enumerate(row):
        if cell in SUPPORTED_FIELDS:
            g_field_2_idx[cell] = idx
    
    if len(g_field_2_idx) != len(SUPPORTED_FIELDS):
        error_exit(f"did not find all of required fields in heading row, given:{g_field_2_idx.keys()}, expected{SUPPORTED_FIELDS}")

    for key, idx in g_field_2_idx.items():
        print(f" key '{key:20}' at idx '{idx}'")
    
    print(f'build entry for fields which require to calculate frequency: {FEQ_FIELDS}')
    for field in FEQ_FIELDS:
        g_f_2_kw_map_feq[field] = {}

def proc_one_row(row: Tuple):
    #should handle by keyword
    for field in FEQ_FIELDS:
        f_cell = row[ g_field_2_idx[field] ]
        if not f_cell:
            raise Exception(f"empty '{field}'")
        
        feq_entry = g_f_2_kw_map_feq[field]
        kw_list = [k.strip() for k in f_cell.split(';')]
        for kw in kw_list:
            lkw = kw.lower()
            if lkw not in feq_entry.keys():
                feq_entry[lkw] = 1
            else:
                feq_entry[lkw] += 1

def save_keyword_feq_to_result_sheet(wb):
    for field in FEQ_FIELDS:
        result_sheet = wb.create_sheet(f"Result {field}")

        for pair in g_f_2_kw_map_feq[field].items():
            result_sheet.append(pair)

    result_sheet = wb.create_sheet("Warnings")
    for warn in g_warnings:
        result_sheet.append((warn,))
    try:
        print(f"update result to result sheet")
        wb.save(FILE_NAME)
    except Exception as e:
        print(f"there is something wrong when updating same file. wrtie result to another file named 'planB.xlsx'")
        wb.save('planB.xlsx')
    #wb.save(FILE_NAME)

'''
Requirement:
    1. filename
    2. to-be-analyzed sheet should named 'tb1'
'''
#read xls
#TODO: handle file extention error
FILE_NAME = "C:\\Users\\qoojo\\OneDrive\\文件\\github\\miniUtil\\cal_kaywords\\full.xlsx"
print(f"reading xlsx:{FILE_NAME}")

try:
    wb = openpyxl.load_workbook(FILE_NAME)
    print("reading workbook successfully")
except PermissionError as e:
    error_exit(f"You probabaly opened {FILE_NAME} by MS Excel. ErrCode:{repr(e)}. Please close other programs which open FILE_NAME.")
except FileNotFoundError as e:
    error_exit(f"Your FILE_NAME '{FILE_NAME}' does not exist. ErrCode:{repr(e)}. Please correctify your FILE_NAME")
except Exception as e:
    print(f"e type: {type(e)}")
    error_exit(f"hit unhandled exception when loading FILE_NAME '{FILE_NAME}'. ErrCode:{repr(e)}")

sheet_names = [f'Result {field}' for field in FEQ_FIELDS]
sheet_names.append('Warnings')
dsn = [name for name in sheet_names if name in wb.sheetnames]
for name in dsn:
    pwarn(f"previous result sheet '{name}' will be removed")
    wb.remove(wb[name])
# wb.save(FILE_NAME)

print(f"ATTENTION ! only sheet named 'tb1' will be analyzed")
tb = wb['tb1']

for row_idx, row in enumerate(tb.iter_rows(values_only=True), start=1):
    if row_idx == 1:
        key_2_idx = verify_head_row_and_build_global_var(row)
        continue
    try:
        #print(row)
        proc_one_row(row)
    except Exception as e:
        pwarn(f'row {row_idx} has unexpected format. ErrCode:{repr(e)}')

print(f"parsing done. hit {len(g_warnings)} warnings")
save_keyword_feq_to_result_sheet(wb)
wb.close()
print("procedure complete. ^_^")